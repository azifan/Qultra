import os
import re
import platform

from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from PyQt6.QtWidgets import QWidget, QApplication, QFileDialog

from src.CeusTool3d.exportData_ui import Ui_exportData

system = platform.system()


class ExportDataGUI(Ui_exportData, QWidget):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.setLayout(self.fullScreenLayout)

        if system == "Windows":
            self.imageSelectionLabelSidebar.setStyleSheet(
                """QLabel {
                font-size: 18px;
                color: rgb(255, 255, 255);
                background-color: rgba(255, 255, 255, 0);
                border: 0px;
                font-weight: bold;
            }"""
            )
            self.imageLabel.setStyleSheet(
                """QLabel {
                font-size: 13px;
                color: rgb(255, 255, 255);
                background-color: rgba(255, 255, 255, 0);
                border: 0px;
                font-weight: bold;
            }"""
            )
            self.imagePathInput.setStyleSheet(
                """QLabel {
                font-size: 11px;
                color: rgb(255, 255, 255);
                background-color: rgba(255, 255, 255, 0);
                border: 0px;
            }"""
            )
            self.roiSidebarLabel.setStyleSheet(
                """QLabel {
                font-size: 18px;
                color: rgb(255, 255, 255);
                background-color: rgba(255, 255, 255, 0);
                border: 0px;
                font-weight: bold;
            }"""
            )
            self.analysisParamsLabel.setStyleSheet(
                """QLabel {
                font-size: 18px;
                color: rgb(255, 255, 255);
                background-color: rgba(255, 255, 255, 0);
                border: 0px;
                font-weight:bold;
            }"""
            )
            self.ticAnalysisLabel.setStyleSheet(
                """QLabel {
                font-size: 18px;
                color: rgb(255, 255, 255);
                background-color: rgba(255, 255, 255, 0);
                border: 0px;
                font-weight: bold;
            }"""
            )
            self.rfAnalysisLabel.setStyleSheet(
                """QLabel {
                font-size: 18px;
                color: rgb(255, 255, 255);
                background-color: rgba(255, 255, 255, 0);
                border: 0px;
                font-weight: bold;
            }"""
            )

        self.hideAppendToFileLayout()
        self.hideNewFileLayout()
        self.dataSavedSuccessfullyLabel.hide()

        self.dataFrame = None
        self.lastGui = None

        self.newFileOptionButton.clicked.connect(self.newFileOptionSelected)
        self.newFileBackButton.clicked.connect(self.backFromNewFileOption)
        self.appendFileOptionButton.clicked.connect(self.appendOptionSelected)
        self.appendFileBackButton.clicked.connect(self.backFromAppendOption)
        self.backButton.clicked.connect(self.backToAnalysis)
        self.appendFileButton.clicked.connect(self.appendToFile)
        self.createNewFileButton.clicked.connect(self.createNewFile)
        self.chooseNewFolderButton.clicked.connect(self.selectNewFolder)
        self.chooseAppendFileButton.clicked.connect(self.selectExistingFile)
        self.clearNewFolderButton.clicked.connect(self.clearNewFolder)
        self.clearAppendFileButton.clicked.connect(self.clearNewFile)

    def hideApproachSelectionLayout(self):
        self.approachSelectionLayout.setContentsMargins(0, 0, 0, 0)
        self.selectDataLabel.hide(); self.newFileOptionButton.hide()
        self.appendFileOptionButton.hide()

    def showApproachSelectionLayout(self):
        self.approachSelectionLayout.setContentsMargins(30, 0, 30, 200)
        self.selectDataLabel.show(); self.newFileOptionButton.show()
        self.appendFileOptionButton.show()
    
    def hideAppendToFileLayout(self):
        self.appendFileHeading.hide(); self.appendFileLabel.hide()
        self.appendFilePath.hide(); self.chooseAppendFileButton.hide()
        self.clearAppendFileButton.hide(); self.appendFileButton.hide()
        self.appendFileBackButton.hide()

    def showAppendToFileLayout(self):
        self.appendFileHeading.show(); self.appendFileLabel.show()
        self.appendFilePath.show(); self.chooseAppendFileButton.show()
        self.clearAppendFileButton.show(); self.appendFileButton.show()
        self.appendFileBackButton.show()

    def hideNewFileLayout(self):
        self.newFileHeading.hide(); self.newFolderPathLabel.hide()
        self.newFolderPathInput.hide(); self.chooseNewFolderButton.hide()
        self.clearNewFolderButton.hide(); self.newFileNameLabel.hide()
        self.newFileNameInput.hide(); self.fileNameWarningLabel.hide()
        self.fileNameErrorLabel.hide(); self.createNewFileButton.hide()
        self.newFileBackButton.hide()

    def showNewFileLayout(self):
        self.newFileHeading.show(); self.newFolderPathLabel.show()
        self.newFolderPathInput.show(); self.chooseNewFolderButton.show()
        self.clearNewFolderButton.show(); self.newFileNameLabel.show()
        self.newFileNameInput.show(); self.fileNameWarningLabel.show()
        self.createNewFileButton.show(); self.newFileBackButton.show()

    def clearNewFolder(self):
        self.newFolderPathInput.clear()

    def setFilenameDisplays(self, imageName):
        self.imagePathInput.setHidden(False)

        imFile = imageName.split("/")[-1]

        self.imagePathInput.setText(imFile)
        self.inputTextPath = imageName

    def clearNewFile(self):
        self.appendFilePath.clear()

    def selectNewFolder(self):
        folderName = QFileDialog.getExistingDirectory(None, "Select Directory")
        if folderName != "":
            self.newFolderPathInput.setText(folderName)

    def selectExistingFile(self):
        fileName, _ = QFileDialog.getOpenFileName(None, "Open file", filter="*.xlsx")
        if fileName != "":
            self.appendFilePath.setText(fileName)

    def dataSavedSuccessfully(self):
        self.hideAppendToFileLayout()
        self.hideNewFileLayout()
        self.dataSavedSuccessfullyLabel.show()

    def createNewFile(self):
        if os.path.exists(self.newFolderPathInput.text()):
            if not (
                self.newFileNameInput.text().endswith(".xlsx")
                and (self.newFileNameInput.text() != ".xlsx")
                and (not bool(re.search(r"\s", self.newFileNameInput.text())))
            ):
                self.fileNameErrorLabel.show()
                return
            try:
                wb = Workbook()
                ws = wb.active
                for r in dataframe_to_rows(self.dataFrame, index=False, header=True):
                    ws.append(r)
                wb.save(
                    os.path.join(
                        self.newFolderPathInput.text(), self.newFileNameInput.text()
                    )
                )
                wb.close()

                self.dataSavedSuccessfully()
                self.dataFrame = self.dataFrame.iloc[0:0]
            except Exception as e:
                print(str(e))

    def appendToFile(self):
        if os.path.exists(
            self.appendFilePath.text()
        ) and self.appendFilePath.text().endswith(".xlsx"):
            try:
                # Since writes to 'Sheet1', make sure not to change sheet names
                wb = load_workbook(self.appendFilePath.text())
                ws = wb.active
                for r in dataframe_to_rows(self.dataFrame, index=False, header=False):
                    ws.append(r)
                wb.save(self.appendFilePath.text())
                wb.close()

                self.dataSavedSuccessfully()
                self.dataFrame = self.dataFrame.iloc[0:0]
            except Exception as e:
                print(str(e))

    def backToAnalysis(self):
        self.lastGui.show()
        self.lastGui.resize(self.size())
        self.hide()

    def newFileOptionSelected(self):
        self.hideApproachSelectionLayout()
        self.showNewFileLayout()

    def backFromNewFileOption(self):
        self.hideNewFileLayout()
        self.showApproachSelectionLayout()

    def appendOptionSelected(self):
        self.hideApproachSelectionLayout()
        self.showAppendToFileLayout()

    def backFromAppendOption(self):
        self.hideAppendToFileLayout()
        self.showApproachSelectionLayout()

if __name__ == "__main__":
    import sys

    app = QApplication(sys.argv)
    # selectWindow = QWidget()
    ui = ExportDataGUI()
    # ui.selectImage.show()
    ui.show()
    sys.exit(app.exec_())

